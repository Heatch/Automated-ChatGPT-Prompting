from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
import time
import pandas as pd
from openpyxl import load_workbook

# RUN BAT FILE FIRST TO START CHROME WITH REMOTE DEBUGGING ENABLED

# Specify the port you've used for remote debugging
debugging_port = 9222
sheet = "./test-file.xlsx"

# Set up Chrome options
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{debugging_port}")

# Path to your ChromeDriver executable
chrome_driver_path = "./chromedriver.exe"

# Set up ChromeDriver with the options
service = Service(chrome_driver_path)
driver = webdriver.Chrome(service=service, options=chrome_options)
actions = ActionChains(driver)

# Find the textarea and submit button
textarea = driver.find_element(By.XPATH, '//*[@id="prompt-textarea"]')

# Read the prompts from the Excel file
df = pd.read_excel(sheet, engine='openpyxl', dtype=object, header=None)
cellTexts = df.values.tolist()
cellTexts = cellTexts[1:]

# Generate responses for each prompt
startingIndex = 2
generationStatus = False
responses = []
for prompt in cellTexts:
    textarea.send_keys(prompt[0])
    actions.send_keys(Keys.ENTER)
    actions.perform()
    while not generationStatus:
        try:
            find_test = driver.find_element(By.XPATH, f'/html/body/div[1]/div[2]/main/div[1]/div[1]/div/div/div/div/article[{startingIndex}]/div/div/div[2]/div/div[2]/div/div/span[1]/button/span')
            generationStatus = True
        except:
            time.sleep(1)
    response = driver.find_element(By.XPATH, f'/html/body/div[1]/div[2]/main/div[1]/div[1]/div/div/div/div/article[{startingIndex}]/div/div/div[2]/div/div[1]/div/div/div')
    responses.append(response.text)
    startingIndex += 2
    generationStatus = False

# Write the responses back to the Excel file
wb = load_workbook(filename=sheet)
ws = wb['Sheet1']

for i in range(len(responses)):
    ws.cell(row=i+2, column=2).value = responses[i]

wb.save(filename=sheet)
